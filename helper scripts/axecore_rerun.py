import os
import json
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from axe_selenium_python import Axe
from openpyxl import load_workbook
from openpyxl.styles import Font

# === CONFIG ===
FIX_DATA_XLSX = "results_cot_rag_generated.xlsx"
OUTPUT_DIR = "axe_wrapped_html_reports_merged"
os.makedirs(OUTPUT_DIR, exist_ok=True)

AXE_JSON_DIR = os.path.join(OUTPUT_DIR, "axe_json")
os.makedirs(AXE_JSON_DIR, exist_ok=True)

HTML_TEMP_DIR = os.path.join(OUTPUT_DIR, "html_temp_for_scan")
os.makedirs(HTML_TEMP_DIR, exist_ok=True)

# === Load Fix Data ===
df_fixes = pd.read_excel(FIX_DATA_XLSX)

# === Headless Chrome Setup ===
options = webdriver.ChromeOptions()
options.add_argument("--headless")
driver = webdriver.Chrome(options=options)

results = []

for idx, row in df_fixes.iterrows():
    html_path = row.get("html_file_path", "")
    file_name = row.get("file", "")  # <--- added
    if not os.path.isfile(html_path):
        continue

    for fix_type in ["cot", "rag"]:
        html_fix = row.get(f"{fix_type}_response", "")
        if not html_fix or "<" not in html_fix:
            continue

        try:
            with open(html_path, "r", encoding="utf-8") as f:
                original_html = f.read()

            soup = BeautifulSoup(original_html, "html.parser")
            old_fragment = row.get("html", "")
            original_tag = BeautifulSoup(old_fragment, "html.parser").find()
            match_tag = None

            if original_tag:
                candidates = soup.find_all(original_tag.name)
                for tag in candidates:
                    if tag.get("id") == original_tag.get("id"):
                        match_tag = tag
                        break
                if not match_tag and candidates:
                    match_tag = candidates[0]

            if match_tag:
                wrapped_section = soup.new_tag("section", **{"data-source": fix_type})
                wrapped_section.append(BeautifulSoup(html_fix, "html.parser"))
                match_tag.replace_with(wrapped_section)
            else:
                fallback = soup.new_tag("section", **{"data-source": fix_type})
                fallback.append(BeautifulSoup(html_fix, "html.parser"))
                soup.body.append(fallback)

            full_html = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>Fix {idx} {fix_type.upper()}</title></head>
<body><main id="axe-scan-target">{str(soup)}</main></body>
</html>"""

            html_file_path = os.path.join(HTML_TEMP_DIR, f"fix_{idx}_{fix_type}.html")
            with open(html_file_path, "w", encoding="utf-8") as f:
                f.write(full_html)

            driver.get("file://" + os.path.abspath(html_file_path))
            axe = Axe(driver)
            axe.inject()
            axe_result = axe.run("main#axe-scan-target")

            violations = axe_result.get("violations", [])
            incomplete = axe_result.get("incomplete", [])
            is_pass = not violations and not incomplete

            violation_descriptions = "; ".join([v.get("description", "") for v in violations])

            axe_json_path = os.path.join(AXE_JSON_DIR, f"{idx}_{fix_type}_axe.json")
            with open(axe_json_path, "w", encoding="utf-8") as f:
                json.dump(axe_result, f, indent=2, ensure_ascii=False)

            evaluation = "✅ Pass" if is_pass else ("⚠️ New Violation" if violations else "❌ Still Failing")

            results.append({
                "index": idx,
                "file": file_name,  # <--- added
                "fix_type": fix_type.upper(),
                "rule_id": row.get("rule_id", ""),
                "evaluation": evaluation,
                "violation_count": len(violations),
                "incomplete_count": len(incomplete),
                "violation_descriptions": violation_descriptions,
                "cot_response": row.get("cot_response", ""),
                "rag_response": row.get("rag_response", ""),
                "axe_json_path": axe_json_path
            })

        except Exception as e:
            results.append({
                "index": idx,
                "file": file_name,  # <--- added
                "fix_type": fix_type.upper(),
                "rule_id": row.get("rule_id", ""),
                "evaluation": f"❌ Error: {e}",
                "violation_count": "N/A",
                "incomplete_count": "N/A",
                "violation_descriptions": str(e),
                "cot_response": row.get("cot_response", ""),
                "rag_response": row.get("rag_response", ""),
                "axe_json_path": "N/A"
            })

driver.quit()

# === Save Outputs ===
summary_df = pd.DataFrame(results)
summary_df.to_csv(os.path.join(OUTPUT_DIR, "axe_evaluation_summary.csv"), index=False)

xlsx_path = os.path.join(OUTPUT_DIR, "axe_evaluation_summary.xlsx")
summary_df.to_excel(xlsx_path, index=False)

wb = load_workbook(xlsx_path)
ws = wb.active
for row in ws.iter_rows(min_row=2):
    if str(row[4].value).startswith("❌") or str(row[4].value).startswith("⚠️"):
        for cell in row:
            cell.font = Font(bold=True)
wb.save(xlsx_path)

print(f"\n✅ Axe Evaluation Complete. Results saved to '{OUTPUT_DIR}'")
